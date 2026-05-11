"""
Structured ESOP data extractor + Excel exporter.
Excel format exactly matches CEINSYS / Wanbury / Almondz reference images.
"""
import json, time
import anthropic
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

from config import ANTHROPIC_API_KEY
from pdf_parser import get_esop_text, get_full_text, MAX_PAGES_STANDARD, MAX_PAGES_EXPANDED
from schema import (
    SCHEME_FIELDS_ORDERED, ESOP_FIELDS, FIELD_LABELS,
    KMP_FIELDS, KMP_FIELD_LABELS,
    HIGHLIGHT_ROWS, SECTION_ROWS,
    SCHEME_EXTRACTION_PROMPT, KMP_EXTRACTION_PROMPT,
)

client = anthropic.Anthropic(api_key=ANTHROPIC_API_KEY)

# ── Colours (matching reference images) ──────────────────────────────────────
C_DARK_BLUE   = "1F4E79"
C_MED_BLUE    = "2E75B6"
C_LIGHT_BLUE  = "D6E4F0"
C_GOLD        = "FFC000"   # yellow highlight rows
C_ORANGE      = "F4B942"   # section dividers for % metrics
C_GREEN_LIGHT = "E2EFDA"
C_GREY        = "F2F2F2"
C_WHITE       = "FFFFFF"
C_DARK_TEXT   = "000000"

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color=C_DARK_TEXT, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic)

def _border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
RIGHT  = Alignment(horizontal="right",  vertical="center")

# ── Claude helpers ────────────────────────────────────────────────────────────

def _call(kwargs, retries=5):
    for attempt in range(retries):
        try:
            return client.messages.create(**kwargs)
        except anthropic.RateLimitError:
            wait = 30 * (attempt + 1)
            print(f"  Rate limit — waiting {wait}s (attempt {attempt+1}/{retries})...")
            time.sleep(wait)
        except anthropic.APIStatusError as e:
            if e.status_code == 529 or "overloaded" in str(e).lower():
                wait = 15 * (attempt + 1)
                print(f"  API overloaded — waiting {wait}s (attempt {attempt+1}/{retries})...")
                time.sleep(wait)
            else:
                raise
        except Exception as e:
            if attempt < retries - 1:
                wait = 10 * (attempt + 1)
                print(f"  API error ({e}) — retrying in {wait}s...")
                time.sleep(wait)
            else:
                raise
    raise RuntimeError("API retries exhausted")


def _parse_json(raw: str, label: str) -> list:
    raw = raw.strip()
    if "```" in raw:
        raw = raw.split("```")[1]
        if raw.startswith("json"): raw = raw[4:]
    raw = raw.strip()
    try:
        return json.loads(raw)
    except json.JSONDecodeError as e:
        print(f"  JSON error in {label}: {e} | snippet: {raw[:150]}")
        return []


def _normalize_scheme_name(name: str) -> str:
    import re
    n = (name or "Unknown").lower().strip()
    year_m = re.search(r"\b(19|20)\d{2}\b", n)
    year   = year_m.group() if year_m else ""

    if "rsu" in n and year:                    return f"{year} RSU"
    if ("esop" in n or "esos" in n) and year:  return f"{year} ESOP"
    if "stock incentive compensation" in n and year: return f"{year} RSU"  # Infosys SICP = RSU umbrella
    if ("expanded stock ownership" in n or " esp" in n) and year: return f"{year} ESP"
    if ("stock option" in n or "option plan" in n) and year: return f"{year} ESOP"
    if year:
        # Try to find a type keyword
        if "restricted" in n:   return f"{year} RSU"
        if "performanc" in n:   return f"{year} PSU"
        if "appreciation" in n: return f"{year} SAR"
        return f"{year} Plan"
    return name[:30].strip()


def _compute_pct_fields(scheme: dict) -> dict:
    """Compute / fix % fields and paid_up_capital."""
    paid_up = scheme.get("paid_up_capital")

    # ── Attempt to back-calculate correct share count from stated % + pool ────
    # Claude sometimes misparses Indian comma notation (e.g. 4,14,85,60,440 → 41.5B
    # instead of 4.15B).  If the PDF already states dilution_pct AND pool_approved,
    # we can recover the true share count which is much more reliable.
    stated_dilution = scheme.get("dilution_pct")
    pool = scheme.get("pool_approved")
    if stated_dilution and pool and float(stated_dilution) > 0:
        try:
            back_calc = float(pool) / float(stated_dilution)
            # Accept if in reasonable range: 10M – 50B shares
            if 1e7 <= back_calc <= 5e10:
                paid_up = back_calc
                scheme["paid_up_capital"] = round(back_calc)
        except (TypeError, ValueError, ZeroDivisionError):
            pass

    # ── Validate / reject the raw extracted paid_up_capital if still bad ─────
    if paid_up is None:
        return scheme
    try:
        paid_up_f = float(paid_up)
    except (TypeError, ValueError):
        scheme["paid_up_capital"] = None
        return scheme

    # Discard implausible values (rupee amounts, garbage, etc.)
    if paid_up_f < 5_000_000 or paid_up_f > 50_000_000_000:
        scheme["paid_up_capital"] = None
        return scheme

    def pct(num):
        if num is not None:
            try: return round(float(num) / paid_up_f, 6)
            except: return None
        return None

    if scheme.get("dilution_pct") is None:
        scheme["dilution_pct"] = pct(scheme.get("pool_approved"))
    if scheme.get("ownership_pct") is None:
        scheme["ownership_pct"] = pct(scheme.get("options_outstanding_end"))
    if scheme.get("overhang_pct") is None:
        scheme["overhang_pct"] = pct(scheme.get("options_outstanding_end"))
    if scheme.get("burn_rate_pct") is None:
        scheme["burn_rate_pct"] = pct(scheme.get("options_granted"))
    return scheme


# ── Extraction helpers ────────────────────────────────────────────────────────

def _data_richness(schemes: list[dict]) -> float:
    """Return 0–1 fraction of non-null numeric fields across all schemes."""
    if not schemes:
        return 0.0
    numeric_keys = [
        "options_outstanding_beginning", "options_granted", "options_vested",
        "options_exercised", "options_lapsed", "options_outstanding_end",
        "pool_approved", "paid_up_capital",
    ]
    total = len(schemes) * len(numeric_keys)
    filled = sum(1 for s in schemes for k in numeric_keys if s.get(k) is not None)
    return filled / max(total, 1)


def _call_claude_extract(prompt: str, max_tokens: int = 4096) -> str:
    resp = _call(dict(
        model="claude-sonnet-4-6",
        max_tokens=max_tokens,
        messages=[{"role": "user", "content": prompt}],
    ))
    return resp.content[0].text


# ── Extraction ────────────────────────────────────────────────────────────────

def extract_scheme_data(company, year, pdf_path, url_info) -> list[dict]:
    """
    Multi-attempt extraction with progressive fallbacks:
    1. Standard: top-15 ESOP pages
    2. Expanded: top-30 ESOP pages (if first attempt yields poor data)
    3. Full-doc: wider text scan (if company uses unusual ESOP terminology)
    """
    pdf_url = (url_info or {}).get("pdf_url", "N/A")

    def _extract_and_call(esop_text: str, label: str) -> list[dict]:
        if not esop_text:
            return []
        prompt = SCHEME_EXTRACTION_PROMPT.format(
            company=company, year=year, pdf_url=pdf_url, esop_text=esop_text)
        print(f"  Extracting schemes FY{year} ({label})...")
        raw = _call_claude_extract(prompt)
        schemes = _parse_json(raw, f"FY{year} {label}")
        for s in schemes:
            s["source_pdf_url"] = pdf_url
            s = _compute_pct_fields(s)
        return schemes

    # ── Attempt 1: standard (top 15 pages) ───────────────────────────────────
    esop_text, _ = get_esop_text(pdf_path, max_pages=MAX_PAGES_STANDARD)
    schemes = _extract_and_call(esop_text, "standard")

    # ── Attempt 2: expanded (top 30 pages) if data is thin ───────────────────
    if _data_richness(schemes) < 0.25:
        print(f"  Data thin ({_data_richness(schemes):.0%}) — retrying with 30 pages...")
        esop_text_wide, _ = get_esop_text(pdf_path, max_pages=MAX_PAGES_EXPANDED)
        if esop_text_wide and esop_text_wide != esop_text:
            schemes2 = _extract_and_call(esop_text_wide, "expanded")
            if _data_richness(schemes2) > _data_richness(schemes):
                schemes = schemes2

    # ── Attempt 3: full-doc scan if still empty ───────────────────────────────
    # Send up to 40,000 chars of the full document to catch unusual ESOP formats
    if not schemes or _data_richness(schemes) < 0.1:
        print(f"  Still no data — trying full-document scan...")
        full_text = get_full_text(pdf_path)
        if full_text:
            # Take first 40K chars (intro + notes section most likely to have ESOP)
            chunk = full_text[:40000]
            schemes3 = _extract_and_call(chunk, "full-doc")
            if _data_richness(schemes3) > _data_richness(schemes):
                schemes = schemes3

    print(f"  FY{year}: {len(schemes)} scheme(s) | richness={_data_richness(schemes):.0%}")
    return schemes


def extract_kmp_data(company, year, pdf_path, pdf_url="") -> list[dict]:
    esop_text, _ = get_esop_text(pdf_path, max_pages=MAX_PAGES_STANDARD)
    if not esop_text:
        return []
    prompt = KMP_EXTRACTION_PROMPT.format(
        company=company, year=year, pdf_url=pdf_url, esop_text=esop_text)
    print(f"  Extracting KMP FY{year}...")
    raw    = _call_claude_extract(prompt, max_tokens=2048)
    kmps   = _parse_json(raw, f"FY{year} KMP")
    fy_label = f"FY{year-1}-{str(year)[-2:]}"
    for k in kmps:
        k["fiscal_year"] = fy_label
        k["document"]    = f"AR {fy_label}.pdf"
    print(f"  FY{year}: {len(kmps)} KMP record(s)")
    return kmps


def extract_all_years(company, pdf_paths, url_map) -> tuple[dict, list, dict]:
    scheme_data, kmp_data = {}, []
    esop_texts: dict[int, str] = {}   # raw ESOP section text per year

    for year, pdf_path in sorted(pdf_paths.items()):
        print(f"\n── FY{year} ──────────────────────")
        url_info = url_map.get(str(year)) or url_map.get(year, {})
        pdf_url  = (url_info or {}).get("pdf_url", "")

        # Capture raw ESOP text for the Statements sheet
        raw_text, _ = get_esop_text(pdf_path, max_pages=MAX_PAGES_EXPANDED)
        if raw_text:
            esop_texts[year] = raw_text

        schemes = extract_scheme_data(company, year, pdf_path, url_info)
        for s in schemes:
            key = _normalize_scheme_name(s.get("scheme_name") or "Unknown")
            if key not in scheme_data: scheme_data[key] = {}
            scheme_data[key][year] = s

        kmps = extract_kmp_data(company, year, pdf_path, pdf_url)
        kmp_data.extend(kmps)

    # ── Propagate paid_up_capital across years within each scheme ─────────────
    for key, year_data in scheme_data.items():
        known_pucs = {
            yr: s["paid_up_capital"]
            for yr, s in year_data.items()
            if s.get("paid_up_capital") and float(s["paid_up_capital"]) > 5e6
        }
        if not known_pucs:
            continue
        ref_puc = sorted(known_pucs.items())[-1][1]
        for yr, s in year_data.items():
            if not s.get("paid_up_capital"):
                s["paid_up_capital"] = ref_puc
                s = _compute_pct_fields(s)
                year_data[yr] = s

    return scheme_data, kmp_data, esop_texts


# ── Excel export ──────────────────────────────────────────────────────────────

def export_to_excel(company, scheme_data, kmp_data, years, output_path, esop_texts=None):
    wb = Workbook()
    wb.remove(wb.active)

    sorted_years = sorted(years)
    fy_labels    = [f"FY {y-1}-{str(y)[-2:]}" for y in sorted_years]

    # Summary sheet first (always sheet 0)
    _build_summary_sheet(wb, company, scheme_data, kmp_data, sorted_years, fy_labels)

    _build_fy_wise(wb, company, scheme_data, sorted_years, fy_labels)

    for scheme_name, year_data in scheme_data.items():
        _build_scheme_sheet(wb, company, scheme_name, year_data, sorted_years, fy_labels)

    if kmp_data:
        _build_kmp_sheet(wb, kmp_data)

    # ESOP Statements sheet last
    if esop_texts:
        _build_statements_sheet(wb, company, esop_texts)

    wb.save(output_path)
    print(f"\nExcel saved → {output_path}")


# ── Scheme sheet (exact format from reference images) ─────────────────────────

def _build_scheme_sheet(wb, company, scheme_name, year_data, years, fy_labels):
    safe = scheme_name[:31].replace("/", "-").replace("\\", "-").replace(":", "-")
    ws   = wb.create_sheet(title=safe)

    n_years   = len(years)
    last_col  = get_column_letter(1 + n_years)
    brdr      = _border()

    # ── Row 1: company (col A) + scheme name (merged B→last) ─────────────────
    ws["A1"] = company
    ws["A1"].font  = _font(bold=True, color=C_WHITE, size=11)
    ws["A1"].fill  = _fill(C_DARK_BLUE)
    ws["A1"].alignment = CENTER

    ws.merge_cells(f"B1:{last_col}1")
    ws["B1"] = scheme_name
    ws["B1"].font  = _font(bold=True, color=C_WHITE, size=11)
    ws["B1"].fill  = _fill(C_MED_BLUE)
    ws["B1"].alignment = CENTER

    # ── Row 2: column headers ─────────────────────────────────────────────────
    ws["A2"] = "Data Field"
    ws["A2"].font  = _font(bold=True, color=C_WHITE)
    ws["A2"].fill  = _fill(C_MED_BLUE)
    ws["A2"].alignment = LEFT
    ws["A2"].border = brdr

    # Highlight the most recent year column in gold (like reference images)
    latest_year_col = 1 + n_years

    for col_i, (year, label) in enumerate(zip(years, fy_labels), start=2):
        c = ws.cell(row=2, column=col_i, value=label)
        c.font      = _font(bold=True, color=C_DARK_TEXT if col_i == latest_year_col else C_WHITE)
        c.fill      = _fill(C_GOLD if col_i == latest_year_col else C_MED_BLUE)
        c.alignment = CENTER
        c.border    = brdr

    # ── Data rows ─────────────────────────────────────────────────────────────
    row_i = 3
    current_section = None

    for field_key, field_label in SCHEME_FIELDS_ORDERED:

        # Blank separator row
        if field_key is None:
            ws.row_dimensions[row_i].height = 6
            row_i += 1
            continue

        # Section divider row
        if field_key in SECTION_ROWS:
            section_label = SECTION_ROWS[field_key]
            if section_label != current_section:
                current_section = section_label
                ws.merge_cells(f"A{row_i}:{last_col}{row_i}")
                sec_cell = ws.cell(row=row_i, column=1, value=section_label)
                sec_cell.font  = _font(bold=True, color=C_WHITE, size=9)
                # Orange for % metrics section, blue for others
                sec_fill = C_ORANGE if "%" in section_label else C_MED_BLUE
                sec_cell.fill  = _fill(sec_fill)
                sec_cell.alignment = LEFT
                ws.row_dimensions[row_i].height = 14
                row_i += 1

        # Label cell
        is_highlight = field_key in HIGHLIGHT_ROWS
        label_fill   = _fill(C_GOLD) if is_highlight else _fill(C_GREY)
        label_font   = _font(bold=True) if is_highlight else _font(bold=False)

        label_cell = ws.cell(row=row_i, column=1, value=field_label)
        label_cell.font      = label_font
        label_cell.fill      = label_fill
        label_cell.alignment = LEFT
        label_cell.border    = brdr

        # Data cells (one per year)
        for col_i, year in enumerate(years, start=2):
            val  = (year_data.get(year) or {}).get(field_key)
            cell = ws.cell(row=row_i, column=col_i, value=val)
            cell.alignment = CENTER
            cell.border    = brdr
            cell.fill      = _fill(C_GOLD) if is_highlight else _fill(C_WHITE)

            # Format % fields
            if field_key in ("dilution_pct", "ownership_pct", "overhang_pct", "burn_rate_pct"):
                if val is not None:
                    cell.number_format = "0.00%"
                    cell.font = _font(bold=True, color="C00000")  # red-ish for % metrics

            # URL styling
            if field_key == "source_pdf_url" and val and val != "N/A":
                cell.font = _font(color="0563C1", italic=True)
                cell.value = val

        ws.row_dimensions[row_i].height = 16
        row_i += 1

    # ── Column widths ─────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 46
    for i in range(2, 2 + n_years):
        ws.column_dimensions[get_column_letter(i)].width = 16

    ws.row_dimensions[1].height = 22
    ws.row_dimensions[2].height = 26
    ws.freeze_panes = "B3"


# ── FY wise summary sheet ─────────────────────────────────────────────────────

def _build_fy_wise(wb, company, scheme_data, years, fy_labels):
    ws = wb.create_sheet("FY wise", 0)
    brdr = _border()
    n_years  = len(years)
    last_col = get_column_letter(1 + n_years)

    # Title
    ws.merge_cells(f"A1:{last_col}1")
    ws["A1"] = f"{company} — ESOP Summary (FY wise)"
    ws["A1"].font      = _font(bold=True, color=C_WHITE, size=12)
    ws["A1"].fill      = _fill(C_DARK_BLUE)
    ws["A1"].alignment = CENTER

    # Summary fields shown in FY wise
    SUMMARY_FIELDS = [
        ("options_granted",           "Option Granted during the Year"),
        ("options_outstanding_end",   "Options Outstanding at End of Year"),
        ("options_exercised",         "Options Exercised"),
        ("options_lapsed",            "Options Lapsed"),
        ("pool_approved",             "Pool Approved"),
        ("pool_balance",              "Pool Balance"),
        ("paid_up_capital",           "Paid up capital on 31st Mar"),
        ("dilution_pct",              "% Dilution"),
        ("ownership_pct",             "% in Ownership"),
        ("overhang_pct",              "% Option Overhang"),
        ("burn_rate_pct",             "Burn Rate"),
        ("weighted_avg_fair_value",   "Weighted Avg Fair Value"),
        ("stock_price_end_of_year",   "Stock Price at End of Year"),
        ("esop_cost",                 "ESOP Cost / Expense (Rs)"),
        ("wealth_creation_pretax",    "Wealth Creation (Pre-tax)"),
    ]

    row = 2
    for scheme_name, year_data in scheme_data.items():

        # Scheme header
        ws.merge_cells(f"A{row}:{last_col}{row}")
        hdr = ws.cell(row=row, column=1, value=scheme_name)
        hdr.font      = _font(bold=True, color=C_WHITE, size=10)
        hdr.fill      = _fill(C_MED_BLUE)
        hdr.alignment = LEFT
        ws.row_dimensions[row].height = 18
        row += 1

        # Year header
        ws.cell(row=row, column=1, value="Field").font  = _font(bold=True)
        ws.cell(row=row, column=1).fill = _fill(C_LIGHT_BLUE)
        ws.cell(row=row, column=1).border = brdr
        ws.cell(row=row, column=1).alignment = LEFT

        for ci, label in enumerate(fy_labels, start=2):
            c = ws.cell(row=row, column=ci, value=label)
            c.font = _font(bold=True, color=C_DARK_TEXT)
            c.fill = _fill(C_LIGHT_BLUE)
            c.alignment = CENTER
            c.border = brdr
        row += 1

        # Data
        for field_key, field_label in SUMMARY_FIELDS:
            is_pct     = "pct" in field_key
            is_special = field_key in ("dilution_pct", "ownership_pct", "overhang_pct", "burn_rate_pct")
            row_fill   = _fill(C_GOLD if is_special else C_WHITE)

            lc = ws.cell(row=row, column=1, value=field_label)
            lc.font      = _font(bold=is_special)
            lc.fill      = _fill(C_GOLD if is_special else C_GREY)
            lc.alignment = LEFT
            lc.border    = brdr

            for ci, year in enumerate(years, start=2):
                val  = (year_data.get(year) or {}).get(field_key)
                cell = ws.cell(row=row, column=ci, value=val)
                cell.alignment = CENTER
                cell.border    = brdr
                cell.fill      = row_fill
                if is_pct and val is not None:
                    cell.number_format = "0.00%"
                    cell.font = _font(bold=True, color="C00000")
            row += 1

        row += 1  # spacer between schemes

    ws.column_dimensions["A"].width = 42
    for i in range(2, 2 + n_years):
        ws.column_dimensions[get_column_letter(i)].width = 16
    ws.freeze_panes = "B3"


# ── KMP ESOPs sheet ──────────────────────────────────────────────────────────

def _build_kmp_sheet(wb, kmp_data):
    ws = wb.create_sheet("KMP ESOPs")
    brdr = _border()
    headers = [KMP_FIELD_LABELS[f] for f in KMP_FIELDS]

    # Header row
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = _font(bold=True, color=C_WHITE)
        c.fill      = _fill(C_DARK_BLUE)
        c.alignment = CENTER
        c.border    = brdr

    # Data rows with alternating year shading
    prev_fy, shade = None, False
    for ri, rec in enumerate(kmp_data, start=2):
        fy = rec.get("fiscal_year", "")
        if fy != prev_fy:
            shade, prev_fy = not shade, fy
        row_fill = _fill(C_LIGHT_BLUE if shade else C_WHITE)

        for ci, field in enumerate(KMP_FIELDS, start=1):
            c = ws.cell(row=ri, column=ci, value=rec.get(field))
            c.fill      = row_fill
            c.border    = brdr
            c.alignment = LEFT if ci <= 4 else CENTER

    widths = [14, 18, 30, 25, 16, 14, 14, 20, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = "A2"


# ── ESOP explanation generator ────────────────────────────────────────────────

def _generate_esop_explanation(company: str, scheme_data: dict, has_esop: bool,
                                total_granted: int, total_pool: int,
                                dilution_pct, kmp_names: list) -> str:
    """
    Ask Claude to write a plain-English explanation of the company's ESOP program.
    Returns a multi-paragraph string (one paragraph per blank-line-separated section).
    """
    if has_esop:
        scheme_summary_lines = []
        for sname, year_data in scheme_data.items():
            yrs = sorted(year_data.keys())
            if not yrs:
                continue
            d = year_data[yrs[-1]]
            def _n(v):
                try: return int(float(v)) if v is not None else None
                except: return None
            granted_latest = _n(d.get("options_granted"))
            outstanding    = _n(d.get("options_outstanding_end"))
            exercise_price = d.get("weighted_avg_exercise_price") or d.get("exercise_price")
            vesting        = d.get("vesting_period") or d.get("vesting_schedule")
            fair_value     = d.get("weighted_avg_fair_value")
            line = f"  - {sname}: first issued ~FY{yrs[0]-1}-{str(yrs[0])[-2:]}"
            if granted_latest: line += f", latest year grants: {granted_latest:,}"
            if outstanding:    line += f", outstanding: {outstanding:,}"
            if exercise_price: line += f", exercise price: ₹{exercise_price}"
            if vesting:        line += f", vesting: {str(vesting)[:60]}"
            if fair_value:     line += f", fair value/option: ₹{fair_value}"
            scheme_summary_lines.append(line)

        prompt = f"""You are writing a plain-English explanation for a financial report about {company}'s Employee Stock Option (ESOP) program.

The reader may be an analyst, investor, or HR professional who understands business but may not be deeply familiar with ESOP mechanics. Write clearly, avoid jargon, and keep it friendly but professional.

Company: {company}
Has ESOP: Yes
Number of schemes: {len(scheme_data)}
Schemes and key data:
{chr(10).join(scheme_summary_lines)}
Total options granted (all years, all schemes): {total_granted:,}
{"Total pool approved: " + str(f"{total_pool:,}") if total_pool else ""}
{"Dilution %: " + str(f"{dilution_pct*100:.2f}%") if dilution_pct else ""}
{"KMP recipients: " + ", ".join(kmp_names[:6]) if kmp_names else ""}

Write the following sections (use blank lines between sections, NO markdown headers or bullets):

SECTION 1 — What are Employee Stock Options? (2-3 sentences explaining ESOPs simply)

SECTION 2 — About {company}'s ESOP Program (2-3 sentences: when they started, how many schemes, purpose, who benefits)

SECTION 3 — Understanding the Schemes (for each scheme, one short paragraph: what the scheme name means (RSU/ESOP/ESOS/SAR etc.), what the pool size means, how vesting works, what exercise price means, and what employees gain when they exercise)

SECTION 4 — What does this mean for the company? (1-2 sentences on dilution impact, alignment of employee and shareholder interests)

Keep the total response under 400 words. Write in flowing prose, no bullet points, no headers."""

    else:
        prompt = f"""You are writing a plain-English explanation for a financial report about {company}.

This company does not appear to have an Employee Stock Option (ESOP) program based on its annual reports.

The reader may be an analyst, investor, or HR professional. Write clearly and keep it friendly but professional.

Company: {company}
Has ESOP: No

Write the following sections (use blank lines between sections, NO markdown headers or bullets):

SECTION 1 — What are Employee Stock Options? (2-3 sentences explaining ESOPs simply — what they are, why companies offer them)

SECTION 2 — Why some companies don't have ESOPs (2-3 sentences: traditional/capital-intensive industries, PSUs, regulatory reasons, preference for cash bonuses or profit-sharing instead)

SECTION 3 — About {company} (1-2 sentences acknowledging that no ESOP data was found in the annual reports reviewed, and suggesting the reader check the company's HR policies or notes to accounts for any stock-linked incentive schemes)

Keep the total response under 200 words. Write in flowing prose, no bullet points, no headers."""

    try:
        resp = _call(dict(
            model="claude-sonnet-4-6",
            max_tokens=1024,
            messages=[{"role": "user", "content": prompt}],
        ))
        return resp.content[0].text.strip()
    except Exception as e:
        print(f"  ESOP explanation generation failed: {e}")
        if has_esop:
            return (
                f"{company} operates an active ESOP program with {len(scheme_data)} scheme(s).\n\n"
                "Employee Stock Options give employees the right to buy company shares at a fixed price "
                "(exercise price) after a waiting period (vesting). This aligns employee incentives with "
                "shareholder interests and is a common tool to attract and retain talent in listed companies."
            )
        else:
            return (
                f"Based on the annual reports reviewed, {company} does not appear to have an active ESOP program.\n\n"
                "Employee Stock Options (ESOPs) give employees the right to buy company shares at a "
                "predetermined price after a vesting period. Not all companies use ESOPs — some prefer "
                "cash bonuses, profit-sharing, or other compensation structures."
            )


# ── Summary sheet ─────────────────────────────────────────────────────────────

def _build_summary_sheet(wb, company, scheme_data, kmp_data, years, fy_labels):
    """
    First sheet: auto-generated overview of the company's ESOP program.
    Includes a written summary + key metrics table.
    """
    ws   = wb.create_sheet("Summary", 0)
    brdr = _border()

    # ── Title banner ──────────────────────────────────────────────────────────
    ws.merge_cells("A1:D1")
    ws["A1"] = f"{company} — ESOP Program Summary"
    ws["A1"].font      = _font(bold=True, color=C_WHITE, size=13)
    ws["A1"].fill      = _fill(C_DARK_BLUE)
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 28

    # ── Build summary text from structured data ────────────────────────────────
    has_esop    = bool(scheme_data)
    scheme_names = list(scheme_data.keys())
    n_schemes    = len(scheme_names)
    first_year   = min(years) if years else None
    latest_year  = max(years) if years else None

    # Aggregate metrics across all schemes (latest year)
    total_pool       = 0
    total_outstanding = 0
    total_granted_all = 0
    latest_dilution   = None
    latest_ownership  = None
    latest_fair_value = None

    for sname, year_data in scheme_data.items():
        for yr in sorted(year_data.keys()):
            d = year_data[yr]
            def _n(v):
                try: return int(float(v)) if v is not None else 0
                except: return 0
            total_granted_all += _n(d.get("options_granted"))

        if latest_year in year_data:
            ld = year_data[latest_year]
            total_pool        += _n(ld.get("pool_approved")) if ld.get("pool_approved") else 0
            total_outstanding += _n(ld.get("options_outstanding_end"))
            if latest_dilution is None and ld.get("dilution_pct"):
                try: latest_dilution  = float(ld["dilution_pct"])
                except: pass
            if latest_ownership is None and ld.get("ownership_pct"):
                try: latest_ownership = float(ld["ownership_pct"])
                except: pass
            if latest_fair_value is None and ld.get("weighted_avg_fair_value"):
                try: latest_fair_value = float(ld["weighted_avg_fair_value"])
                except: pass

    kmp_names = list({k.get("kmp_name","") for k in kmp_data if k.get("kmp_name")})

    fy_range = ""
    if years:
        if first_year != latest_year:
            fy_range = f"FY{first_year-1}-{str(first_year)[-2:]} to FY{latest_year-1}-{str(latest_year)[-2:]}"
        else:
            fy_range = f"FY{latest_year-1}-{str(latest_year)[-2:]}"

    # ── Generate AI explanation ────────────────────────────────────────────────
    print("  Generating ESOP explanation...")
    explanation_text = _generate_esop_explanation(
        company, scheme_data, has_esop,
        total_granted_all, total_pool, latest_dilution, kmp_names,
    )

    # ── Section helper ─────────────────────────────────────────────────────────
    def _write_section_header(ws, row, title, fill_color=None):
        ws.merge_cells(f"A{row}:D{row}")
        cell = ws.cell(row=row, column=1, value=title)
        cell.font      = _font(bold=True, color=C_WHITE, size=10)
        cell.fill      = _fill(fill_color or C_MED_BLUE)
        cell.alignment = LEFT
        ws.row_dimensions[row].height = 20
        return row + 1

    def _write_text_block(ws, row, text, bg_even=C_GREY, bg_odd=C_WHITE):
        paragraphs = [p.strip() for p in text.split("\n") if p.strip()]
        for i, para in enumerate(paragraphs):
            ws.merge_cells(f"A{row}:D{row}")
            cell = ws.cell(row=row, column=1, value=para)
            cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
            cell.font      = _font(size=10)
            cell.fill      = _fill(bg_even if i % 2 == 0 else bg_odd)
            # Estimate height: ~110 chars per line at col width 120
            lines = max(1, len(para) // 110 + 1)
            ws.row_dimensions[row].height = max(16, lines * 15)
            row += 1
        return row

    # ── Row 2: About this report section header ────────────────────────────────
    row = 2
    row = _write_section_header(ws, row, "About this Report")

    about_lines = [
        f"Company: {company}",
        f"ESOP Status: {'✅ Active ESOP Program' if has_esop else '❌ No ESOP Program Found'}",
        f"Years Analyzed: {len(years)}  ({fy_range})" if fy_range else f"Years Analyzed: {len(years)}",
    ]
    if has_esop:
        about_lines += [
            f"Number of Schemes: {n_schemes}   |   Schemes: {', '.join(scheme_names)}",
            f"Total Options Granted (All Years, All Schemes): {total_granted_all:,}" if total_granted_all else "",
            f"Pool Approved (Latest Year): {total_pool:,}" if total_pool else "",
            f"Options Outstanding (Latest Year-end): {total_outstanding:,}" if total_outstanding else "",
            f"Dilution % (Pool / Paid-up Capital, Latest Year): {latest_dilution*100:.2f}%" if latest_dilution else "",
            f"Ownership % (Outstanding / Paid-up Capital, Latest Year): {latest_ownership*100:.2f}%" if latest_ownership else "",
            f"Avg Fair Value per Option (Latest Year): ₹{latest_fair_value:,.2f}" if latest_fair_value else "",
            f"KMP Grant Recipients: {len(kmp_names)}  ({', '.join(kmp_names[:6])})" if kmp_names else "",
        ]
    about_lines = [l for l in about_lines if l]
    about_lines += ["", "Note: All data extracted from BSE-filed annual reports using AI."]

    for line in about_lines:
        ws.merge_cells(f"A{row}:D{row}")
        is_metric = "%" in line or "Total" in line or "Pool" in line or "Outstanding" in line
        cell = ws.cell(row=row, column=1, value=line)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.font  = _font(bold=is_metric, size=10,
                           color="C00000" if ("%" in line and "Status" not in line) else C_DARK_TEXT)
        cell.fill  = _fill(C_GOLD if is_metric else (C_GREY if row % 2 == 0 else C_WHITE))
        ws.row_dimensions[row].height = 16
        row += 1

    row += 1  # spacer

    # ── AI explanation section ─────────────────────────────────────────────────
    section_title = (
        "Understanding the ESOP Program" if has_esop
        else "What are Employee Stock Options (ESOPs)?"
    )
    row = _write_section_header(ws, row, section_title, fill_color=C_DARK_BLUE)
    row = _write_text_block(ws, row, explanation_text)
    row += 1  # spacer

    ws.column_dimensions["A"].width = 55
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 18


# ── ESOP Statements sheet ─────────────────────────────────────────────────────

# Signals that confirm a block of text is genuinely about ESOPs, not boilerplate
_ESOP_CONFIRM_SIGNALS = [
    "options granted", "options vested", "options exercised", "options lapsed",
    "options outstanding", "options forfeited", "stock option", "stock options",
    "esop", "esos", "restricted stock unit", "rsu", "stock appreciation right",
    "black-scholes", "black scholes", "fair value of option", "weighted average exercise",
    "vesting schedule", "exercise price", "grant date", "option pool",
    "employee stock", "share-based payment", "share based payment",
    "stock-based compensation", "stock based compensation",
    "equity incentive", "equity award", "long-term incentive plan",
    "units granted", "units vested", "units forfeited",
    "pool approved", "dilution", "paid-up capital",
]


def _join_pdf_lines(raw_text: str) -> list[str]:
    """
    Convert raw PDF-extracted text into clean paragraphs.
    PDF extraction splits every line; we need to re-join them into sentences/paragraphs.
    Rule: blank lines mark paragraph breaks; consecutive non-blank lines get joined with space.
    """
    paragraphs, current = [], []
    for line in raw_text.splitlines():
        stripped = line.strip()
        # Page markers become paragraph separators
        if stripped.startswith("[PAGE ") and stripped.endswith("]"):
            if current:
                paragraphs.append(" ".join(current))
                current = []
            paragraphs.append(stripped)  # keep page marker as its own paragraph
        elif stripped == "---":
            if current:
                paragraphs.append(" ".join(current))
                current = []
        elif stripped:
            current.append(stripped)
        else:
            # Blank line → end of paragraph
            if current:
                paragraphs.append(" ".join(current))
                current = []
    if current:
        paragraphs.append(" ".join(current))

    # Remove paragraphs that are purely page markers with no ESOP content
    return [p for p in paragraphs if p]


def _is_esop_block(text: str) -> bool:
    """Return True if the text block contains at least one real ESOP signal."""
    if not text:
        return False
    t = text.lower()
    return any(sig in t for sig in _ESOP_CONFIRM_SIGNALS)


def _build_statements_sheet(wb, company, esop_texts: dict):
    """
    Last sheet: ESOP-specific text from each annual report, organized by year.
    Lines are rejoined into paragraphs; non-ESOP boilerplate pages are excluded.
    """
    ws = wb.create_sheet("ESOP Statements")

    # Title
    ws.merge_cells("A1:B1")
    ws["A1"] = f"{company} — ESOP Disclosures from Annual Reports"
    ws["A1"].font      = _font(bold=True, color=C_WHITE, size=12)
    ws["A1"].fill      = _fill(C_DARK_BLUE)
    ws["A1"].alignment = CENTER
    ws.row_dimensions[1].height = 26

    row = 2
    any_content_written = False

    for year in sorted(esop_texts.keys()):
        fy_label = f"FY {year-1}-{str(year)[-2:]}"
        raw_text  = esop_texts[year] or ""

        # Split into page blocks and keep only those with genuine ESOP signals
        page_blocks = raw_text.split("\n\n---\n\n")
        esop_blocks = [b for b in page_blocks if _is_esop_block(b)]

        if not esop_blocks:
            # No real ESOP content this year — skip the year entirely
            continue

        # Year header
        ws.merge_cells(f"A{row}:B{row}")
        hdr = ws.cell(row=row, column=1, value=f"Annual Report — {fy_label}")
        hdr.font      = _font(bold=True, color=C_WHITE, size=10)
        hdr.fill      = _fill(C_MED_BLUE)
        hdr.alignment = LEFT
        ws.row_dimensions[row].height = 20
        row += 1
        any_content_written = True

        for block in esop_blocks:
            paragraphs = _join_pdf_lines(block)
            shade = False

            for para in paragraphs:
                if not para.strip():
                    continue

                is_page_marker = para.startswith("[PAGE ") and para.endswith("]")
                shade = not shade

                ws.merge_cells(f"A{row}:B{row}")
                cell = ws.cell(row=row, column=1, value=para)
                cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                cell.font      = _font(
                    bold=is_page_marker,
                    color="2E75B6" if is_page_marker else C_DARK_TEXT,
                    size=9,
                )
                cell.fill = _fill(C_LIGHT_BLUE if is_page_marker else (C_GREY if shade else C_WHITE))

                # Row height: estimate ~130 chars per line at col width 120
                line_count = max(1, len(para) // 130 + 1)
                ws.row_dimensions[row].height = max(14, line_count * 14)
                row += 1

            # Blank separator between page blocks
            row += 1

        # Extra spacer between years
        row += 1

    if not any_content_written:
        ws.merge_cells("A2:B2")
        ws["A2"] = "(No ESOP-specific text found in the annual reports for this company)"
        ws["A2"].font      = _font(italic=True, color="7F7F7F")
        ws["A2"].alignment = LEFT

    ws.column_dimensions["A"].width = 120
    ws.column_dimensions["B"].width = 10
    ws.freeze_panes = "A2"
